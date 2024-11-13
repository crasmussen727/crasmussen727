#Calvin Rasmussen
#11/19/2024
#Lab 12
#Notes: I took insipration from Kirby's Dreamland 2 and this youtube video https://www.youtube.com/watch?v=l2vETbof-qI

import openpyxl
from openpyxl.styles import PatternFill

# Initialize a workbook and set up the active sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Kirby Pixel Art"

# Step 1: Set cell dimensions to create a square appearance
cell_width = 2  
cell_height = cell_width * 6  # Approximate to keep cells square
for col in range(1, 20):  
    col_letter = openpyxl.utils.get_column_letter(col)
    sheet.column_dimensions[col_letter].width = cell_width
for row in range(1, 20):  
    sheet.row_dimensions[row].height = cell_height

# Step 2: Define the pixel art pattern using a dictionary
# Format: {color_code: [cell_coordinates]}
Kirby_art_data = {
    "#FFB6C1": ["H2", "I2", "J2", "F3", "G3", "H3", "I3", "J3", "K3", "L3", "E4", "F4", "G4", "H4", "I4", "J4", "K4", "L4", "M4", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5",
                "C6", "D6", "E6", "F6","G6","H6", "I6","K6","M6","C7","D7","E7","F7","G7","H7","I7","K7","M7","N7","O7",
                "B8","C8","D8","E8","F8","G8","H8","I8","K8","M8","N8","O8","B9","C9","D9","E9","F9","I9","J9","K9","L9","O9","C10",
                "D10","F10","G10","H10","I10","J10","K10","L10","M10","O10","D11","F11","G11","H11","I11","J11","L11","M11",
                "F12","G12","H12","I12","J12","K12","L12","H13","I13","J13","K13"],#light pink
    "#ADD8E6": ["J8","L8"],
    "#FF0000": ["D4","E4","F4","G4","H4","I4","J4","K4","L4","M4"],
    "#FF69B4": ["G2", "K2","E3","M3","D4","N5","N6","B7","G9","H9","M9","N9", "B10","E10","N10","C11","O11","E12","M12","F13","G13","L13",
                "D14","E14","M14","N14","C15","D15","E15","F15","G15","L15","M15","N15","O15"],  # Dark pink
    "#000000": ["G1","J6","L6", "H1", "I1", "J1", "K1", "E2", "F2", "L2", "M2", "D3", "N3", "C4", "N4","C5", "O5","B6","O6",
                "A7","J7","L7","P7","A8","P8","A9","P9","A10","P10","B11","E11","K11","N11","P11","C12","D12","N12",
                "O12","D13","E13","M13","N13","C14","F14","G14","H14","I14","J14","K14","L14","O14","B15","H15","I15","J15","K15","P15",
                "C16","D16","E16","F16","G16","K16","L16","M16","N16","O16"],  # Eyes, mouth, and outline
}

# Step 3: Apply colors to the specified cells
for color, cells in Kirby_art_data.items():
    fill = PatternFill(start_color=color[1:], end_color=color[1:], fill_type="solid")
    for cell in cells:
        sheet[cell].fill = fill

# Save the workbook
wb.save("kirby_pixel_art.xlsx")
